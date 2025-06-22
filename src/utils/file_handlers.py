import os
import shutil
import pandas as pd
import logging.config
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import List
from src.core.logger import logging_config


logging.config.dictConfig(logging_config)
logger = logging.getLogger(name="file_handlers")


def load_companies_from_excel(filepath: str):
    """
    Loads a list of companies from an Excel file.
    The file is expected to have a column named "Company".
    """
    df = pd.read_excel(filepath)
    return df["Company"].dropna().astype(str).tolist()


def save_results_to_excel(
    results: dict,
    original_companies: List[str],
    normalized_companies: List[str],
    output_file: str,
):
    """Saves sanctions check results to a color-coded Excel file."""
    logger.info(
        f"Saving results for {len(original_companies)} "
        f"companies to {output_file}"
    )
    data = []
    for original, normalized in zip(original_companies, normalized_companies):
        status = {
            key: "Yes" if normalized in matches else "No"
            for key, matches in results.items()
        }
        matched_lists = [k for k, v in status.items() if v == "Yes"]
        info = (
            f"Sanctions found in — {matched_lists}"
            if matched_lists
            else "No sanctions found"
        )
        data.append({"Company": original, **status, "Sanctions Info": info})
    df_out = pd.DataFrame(data)
    df_out.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    fill_yes = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    fill_no = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column - 1):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = fill_yes
            elif cell.value == "No":
                cell.fill = fill_no
    wb.save(output_file)
    logger.info(f"Results saved to file: {output_file}")


def clean_tmp_folders(folders: list):
    """Deletes all files and subdirectories within the specified folders."""
    logger.info("Clean tmp folders")
    for folder in folders:
        if os.path.exists(folder):
            for filename in os.listdir(folder):
                file_path = os.path.join(folder, filename)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        shutil.rmtree(file_path)
                except Exception as e:
                    logger.warning(f"Could not delete {file_path}: {e}")
