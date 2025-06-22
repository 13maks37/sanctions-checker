import requests
import re
import os
import pandas as pd
import logging.config
import xml.etree.ElementTree as ET
from datetime import datetime
from aiogram import Bot
from aiogram.types import FSInputFile
from pathlib import Path
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz.fuzz import token_set_ratio
from bs4 import BeautifulSoup
from src.core.config import settings
from src.core.logger import logging_config


logging.config.dictConfig(logging_config)
logger = logging.getLogger(name="sanctions_scraper")


def download_file(url: str, filename: Path):
    logger.info(f"Downloading: {url}")
    response = requests.get(url, stream=True)
    if response.status_code == 200:
        filename.write_bytes(response.content)
        logger.info(f"Saved: {filename}")
    else:
        logger.error(f"Error downloading {url}: {response.status_code}")


def normalize_company_name(company_names: str):
    """
    Normalizes the company name by removing the content in parentheses
    """
    normalize_names = []
    for name in company_names:
        name = re.sub(r"\s*\([^()]*\)", "", name).strip()
        name = re.sub(r"\s+", " ", name).strip()
        normalize_names.append(name)
    return normalize_names


def is_similar(a: str, b: str, threshold: float = 85) -> bool:
    ratio = token_set_ratio(a.lower(), b.lower())
    return ratio >= threshold


def load_companies_from_excel(filepath: str) -> List[str]:
    df = pd.read_excel(filepath)
    return df["Company"].dropna().astype(str).tolist()


def search_company_in_csv(
    file: Path,
    companies: List[str],
    source_name: str,
) -> List[str]:
    try:
        df = pd.read_csv(
            file,
            encoding="utf-8",
            low_memory=False,
            header=None,
        )
    except Exception:
        df = pd.read_csv(
            file,
            encoding="latin1",
            low_memory=False,
            header=None,
        )
    if source_name == "OFAC":
        candidates = df[1].astype(str).tolist()
    else:
        candidates = df.astype(str).agg(" ".join, axis=1).tolist()

    normalize_companies = normalize_company_name(companies)
    found = []
    for c in normalize_companies:
        if any(is_similar(c, candidate) for candidate in candidates):
            found.append(c)
    return found


def search_company_in_xml(
    file: Path,
    companies: List[str],
    source_name: str,
) -> List[str]:
    root = ET.parse(file).getroot()

    if source_name == "UK":
        candidates = []
        for name_elem in root.findall(".//Names/Name/Name6"):
            if name_elem.text:
                candidates.append(name_elem.text.strip())
    elif source_name == "UN" or source_name == "UN-SC":
        candidates = []
        for individual in root.findall(".//INDIVIDUAL"):
            first_name = individual.findtext("FIRST_NAME")
            second_name = individual.findtext("SECOND_NAME")
            if first_name:
                candidates.append(first_name.strip())
            if second_name:
                candidates.append(second_name.strip())
            for alias in individual.findall("INDIVIDUAL_ALIAS"):
                alias_name = alias.findtext("ALIAS_NAME")
                if alias_name and alias_name.strip():
                    candidates.append(alias_name.strip())
    else:
        text = ET.tostring(root, encoding="utf-8", method="text").decode(
            "utf-8"
        )
        candidates = [
            line.strip() for line in text.splitlines() if line.strip()
        ]

    normalize_companies = normalize_company_name(companies)
    found = []
    for c in normalize_companies:
        if any(is_similar(c, candidate) for candidate in candidates):
            found.append(c)
    return found


def search_company_in_html(
    file: Path,
    companies: List[str],
    source_name: str,
) -> List[str]:
    text = file.read_text(encoding="utf-8", errors="ignore")
    if source_name == "EU-Tracker":
        soup = BeautifulSoup(text, "html.parser")
        candidates = [a["title"] for a in soup.select("ul li a[title]")]
    else:
        candidates = text.splitlines()
    normalize_companies = normalize_company_name(companies)
    found = []
    for c in normalize_companies:
        if any(is_similar(c, candidate) for candidate in candidates):
            found.append(c)
    return found


def save_results_to_excel(
    results: dict,
    original_companies: List[str],
    normalized_companies: List[str],
    output_file: str,
):
    logger.info(
        f"Saving results for {len(original_companies)} "
        f"companies to {output_file}"
    )
    data = []
    for original, normalized in zip(original_companies, normalized_companies):
        status = {
            key: "Yes" if normalized in matches else "No"
            for key, matches in results.items()
        }
        matched_lists = [k for k, v in status.items() if v == "Yes"]
        info = (
            f"Sanctions found in — {matched_lists}"
            if matched_lists
            else "No sanctions found"
        )
        data.append({"Company": original, **status, "Sanctions Info": info})
    df_out = pd.DataFrame(data)
    df_out.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    fill_yes = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    fill_no = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column - 1):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = fill_yes
            elif cell.value == "No":
                cell.fill = fill_no
    wb.save(output_file)
    logger.info(f"Results saved to file: {output_file}")


async def scrape_sanctions_companies(
    uploaded_file_path: str, chat_id: int, bot: Bot
):
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = (
        f"{settings.TMP_DIR_RESULT}/sanctions_companies_{date_str}.xlsx"
    )
    logger.info("Starting sanctions check process")
    original_companies = load_companies_from_excel(uploaded_file_path)
    logger.info(f"Loaded {len(original_companies)} companies from input file")
    normalized_companies = normalize_company_name(original_companies)
    os.makedirs(settings.TMP_DIR_SCRAPER, exist_ok=True)
    os.makedirs(settings.TMP_DIR_RESULT, exist_ok=True)
    results = {}
    for name, source in settings.SANCTIONS_SOURCES.items():
        url = source["url"]
        ext = source["ext"]
        file_path = Path(f"{settings.TMP_DIR_SCRAPER}/{name}{ext}")
        logger.info(f"Processing {name} sanctions list from {url}")
        try:
            download_file(url, file_path)
            if ext == ".csv":
                matches = search_company_in_csv(
                    file=file_path,
                    companies=normalized_companies,
                    source_name=name,
                )
            elif ext == ".xml":
                matches = search_company_in_xml(
                    file=file_path,
                    companies=normalized_companies,
                    source_name=name,
                )
            else:
                matches = search_company_in_html(
                    file=file_path,
                    companies=normalized_companies,
                    source_name=name,
                )
            results[name] = matches
            logger.info(f"Processed {name}.Found {len(matches)} matches")
        except Exception as e:
            logger.error(f"Failed to process {name}: {e}", exc_info=True)
            results[name] = []
    logger.info("Generating final report...")
    save_results_to_excel(
        results, original_companies, normalized_companies, output_file
    )
    ready_file = FSInputFile(path=output_file)
    await bot.send_document(
        chat_id=chat_id,
        caption="Sanctions check completed",
        document=ready_file,
    )
    logger.info("Results successfully sent to user")
